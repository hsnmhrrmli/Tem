# -*- coding: utf-8 -*-
"""
excel_exporter.py
------------------
Çıxarılmış başlıq və cədvəl məlumatlarını formatlı Excel (.xlsx) faylına yazır.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from invoice_parser import HEADER_FIELDS_ORDER

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='B7B7B7'),
    right=Side(style='thin', color='B7B7B7'),
    top=Side(style='thin', color='B7B7B7'),
    bottom=Side(style='thin', color='B7B7B7'),
)


def export_to_excel(header: dict, rows: list, columns: list, output_path: str):
    wb = Workbook()

    # ---- Sheet 1: Cədvəl (əsas məlumat - fayl açılanda İLK göstərilən vərəq) ----
    ws2 = wb.active
    ws2.title = "Cədvəl"
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER
    ws2.row_dimensions[1].height = 45
    ws2.freeze_panes = "A2"

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if isinstance(value, float):
                cell.number_format = "0.0000"

    # Cəmi (yekun) sətri - sütun 18 (Yekun məbləğ) üzrə avtomatik cəm
    total_row = len(rows) + 2
    ws2.cell(row=total_row, column=1, value="Cəmi").font = LABEL_FONT
    last_col_letter = get_column_letter(len(columns))
    for col_idx in range(6, len(columns) + 1):  # 6..18: ədədi sütunlar
        col_letter = get_column_letter(col_idx)
        formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        c = ws2.cell(row=total_row, column=col_idx, value=formula)
        c.font = LABEL_FONT
        c.number_format = "0.0000"

    # Sütun enləri
    widths = [8, 42, 16, 12, 10, 10, 14, 14, 10, 12, 16, 14, 14, 14, 16, 14, 12, 16]
    for idx, w in enumerate(widths, start=1):
        if idx <= len(columns):
            ws2.column_dimensions[get_column_letter(idx)].width = w

    # ---- Sheet 2: Başlıq məlumatları (əlavə arayış - ikinci vərəq) ----
    ws1 = wb.create_sheet("Başlıq")
    ws1["A1"] = "Elektron qaimə-faktura — başlıq məlumatları"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:B1")

    r = 3
    for field in HEADER_FIELDS_ORDER:
        if field in header:
            ws1.cell(row=r, column=1, value=field).font = LABEL_FONT
            ws1.cell(row=r, column=2, value=header[field])
            r += 1
    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 60

    # Fayl açılanda "Cədvəl" vərəqi aktiv/görünən olsun
    wb.active = wb.sheetnames.index("Cədvəl")

    wb.save(output_path)
